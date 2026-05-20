# backend/modules/report_service.py — nâng cấp với lợi nhuận & tồn kho
from backend.database.connection import get_connection


def get_dashboard() -> dict:
    conn = get_connection()
    cur  = conn.cursor()
    so_hd, doanh_thu = 0, 0.0
    for col in ("ngay_ban", "created_at", "ngay_tao"):
        try:
            cur.execute(f"""
                SELECT COUNT(*), COALESCE(SUM(thanh_toan), 0)
                FROM hoa_don WHERE DATE({col}) = CURDATE()
            """)
            so_hd, doanh_thu = cur.fetchone()
            break
        except Exception:
            continue

    try:
        cur.execute(
            "SELECT COUNT(*) FROM san_pham WHERE COALESCE(trang_thai,1)=1")
        tong_sp = cur.fetchone()[0]
    except Exception:
        cur.execute("SELECT COUNT(*) FROM san_pham")
        tong_sp = cur.fetchone()[0]

    try:
        from backend.modules.config_service import get_int
        nguong = get_int("canh_bao_ton_kho", 10)
    except Exception:
        nguong = 10

    try:
        cur.execute("""
            SELECT COUNT(*) FROM san_pham
            WHERE ton_kho <= %s AND COALESCE(trang_thai,1)=1
        """, (nguong,))
        sap_het = cur.fetchone()[0]
    except Exception:
        cur.execute("SELECT COUNT(*) FROM san_pham WHERE ton_kho<=10")
        sap_het = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM khach_hang")
    tong_kh = cur.fetchone()[0]

    # Tồn kho tổng giá trị
    try:
        cur.execute("""
            SELECT COALESCE(SUM(ton_kho * gia_nhap), 0)
            FROM san_pham WHERE COALESCE(trang_thai,1)=1
        """)
        gia_tri_kho = float(cur.fetchone()[0] or 0)
    except Exception:
        gia_tri_kho = 0.0

    conn.close()
    return {
        "so_hoa_don" : int(so_hd or 0),
        "doanh_thu"  : float(doanh_thu or 0),
        "tong_sp"    : int(tong_sp),
        "sap_het"    : int(sap_het),
        "tong_kh"    : int(tong_kh),
        "gia_tri_kho": gia_tri_kho,
    }


def _date_col() -> str:
    conn = get_connection()
    cur  = conn.cursor()
    cur.execute("""
        SELECT COLUMN_NAME FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='hoa_don'
          AND COLUMN_NAME IN ('ngay_ban','created_at','ngay_tao')
        LIMIT 1
    """)
    row = cur.fetchone()
    conn.close()
    return row[0] if row else "ngay_ban"


def revenue_by_day(tu_ngay: str, den_ngay: str) -> list:
    dc = _date_col()
    conn = get_connection(); cur = conn.cursor()
    cur.execute(f"""
        SELECT DATE({dc}), COUNT(*), COALESCE(SUM(thanh_toan), 0)
        FROM hoa_don
        WHERE DATE({dc}) BETWEEN %s AND %s
        GROUP BY DATE({dc}) ORDER BY DATE({dc})
    """, (tu_ngay, den_ngay))
    result = cur.fetchall()
    conn.close()
    return result


def profit_by_day(tu_ngay: str, den_ngay: str) -> list:
    """Lợi nhuận = doanh thu − giá vốn hàng bán (COGS)."""
    dc = _date_col()
    conn = get_connection(); cur = conn.cursor()
    cur.execute(f"""
        SELECT DATE(hd.{dc}) AS ngay,
               COUNT(DISTINCT hd.ma_hd)        AS so_hd,
               COALESCE(SUM(hd.thanh_toan), 0) AS doanh_thu,
               COALESCE(SUM(ct.so_luong * sp.gia_nhap), 0) AS von,
               COALESCE(SUM(hd.thanh_toan), 0)
                 - COALESCE(SUM(ct.so_luong * sp.gia_nhap), 0) AS loi_nhuan
        FROM hoa_don hd
        JOIN chi_tiet_hoa_don ct ON hd.ma_hd = ct.ma_hd
        JOIN san_pham sp ON ct.ma_sp = sp.ma_sp
        WHERE DATE(hd.{dc}) BETWEEN %s AND %s
        GROUP BY DATE(hd.{dc}) ORDER BY DATE(hd.{dc})
    """, (tu_ngay, den_ngay))
    result = cur.fetchall()
    conn.close()
    return result


def top_products(limit: int = 10) -> list:
    conn = get_connection(); cur = conn.cursor()
    cur.execute("""
        SELECT sp.ten_sp,
               SUM(ct.so_luong)              AS tong_ban,
               SUM(ct.don_gia * ct.so_luong) AS doanh_thu
        FROM chi_tiet_hoa_don ct
        JOIN san_pham sp ON ct.ma_sp = sp.ma_sp
        GROUP BY sp.ten_sp
        ORDER BY tong_ban DESC LIMIT %s
    """, (limit,))
    result = cur.fetchall()
    conn.close()
    return result


def inventory_report() -> list:
    """Báo cáo tồn kho theo loại sản phẩm."""
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT l.ten_loai,
                   COUNT(sp.ma_sp)           AS so_sp,
                   SUM(sp.ton_kho)           AS tong_ton,
                   SUM(sp.ton_kho * sp.gia_nhap) AS gia_tri_nhap,
                   SUM(sp.ton_kho * sp.gia_ban)  AS gia_tri_ban
            FROM san_pham sp
            LEFT JOIN loai_san_pham l ON sp.ma_loai = l.ma_loai
            WHERE COALESCE(sp.trang_thai,1) = 1
            GROUP BY l.ten_loai
            ORDER BY gia_tri_nhap DESC
        """)
        result = cur.fetchall()
    except Exception:
        result = []
    conn.close()
    return result


def low_stock_products() -> list:
    """Sản phẩm cần đặt hàng thêm."""
    try:
        from backend.modules.config_service import get_int
        nguong = get_int("canh_bao_ton_kho", 10)
    except Exception:
        nguong = 10
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT sp.ma_code, sp.ten_sp,
                   COALESCE(l.ten_loai,'—') AS ten_loai,
                   sp.ton_kho,
                   COALESCE(sp.ton_kho_min, %s) AS ton_kho_min,
                   sp.gia_nhap
            FROM san_pham sp
            LEFT JOIN loai_san_pham l ON sp.ma_loai = l.ma_loai
            WHERE sp.ton_kho <= %s
              AND COALESCE(sp.trang_thai,1)=1
            ORDER BY sp.ton_kho ASC
        """, (nguong, nguong))
        result = cur.fetchall()
    except Exception:
        result = []
    conn.close()
    return result


def export_excel(data: list, filepath: str,
                 sheet_name: str = "Doanh thu",
                 headers: list | None = None) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if headers is None:
            headers = ["Ngày", "Số hóa đơn", "Doanh thu (đ)"]

        fill = PatternFill("solid", fgColor="1565C0")
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = fill
            c.alignment = Alignment(horizontal="center")

        for i, row in enumerate(data, 2):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)

        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 22

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Lỗi xuất Excel: {e}")
        return False


def plot_revenue(data: list, mode: str = "revenue"):
    import matplotlib
    matplotlib.use("TkAgg")

    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import numpy as np

    plt.style.use("ggplot")

    ngay = [str(r[0]) for r in data]
    n = len(ngay)

    fig_width = max(9, min(n * 1.2, 16))
    fig_height = 5.2

    fig, ax = plt.subplots(figsize=(fig_width, fig_height))

    x = np.arange(n)

    if mode == "profit" and len(data[0]) >= 5:

        dt = [float(r[2] or 0) for r in data]
        loi = [float(r[4] or 0) for r in data]

        width = 0.32

        bars1 = ax.bar(
            x - width / 2,
            dt,
            width,
            label="Doanh thu",
            color="#1976D2",
            edgecolor="white",
            linewidth=1
        )

        bars2 = ax.bar(
            x + width / 2,
            loi,
            width,
            label="Lợi nhuận",
            color="#2E7D32",
            edgecolor="white",
            linewidth=1
        )

        ymax = max(max(dt), max(loi))
        ymin = min(min(loi), 0)

        ax.set_ylim(ymin * 1.2, ymax * 1.18)

        for bars in [bars1, bars2]:
            for bar in bars:
                val = bar.get_height()

                if abs(val) < 1:
                    continue

                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    val + ymax * 0.015 if val > 0 else val - ymax * 0.05,
                    f"{val/1e6:.1f}M",
                    ha="center",
                    fontsize=8,
                    fontweight="bold"
                )

        ax.legend(
            frameon=False,
            fontsize=10,
            loc="upper right"
        )

        ax.set_title(
            "Doanh thu & Lợi nhuận theo ngày",
            fontsize=16,
            fontweight="bold",
            pad=18
        )

    else:

        dt = [float(r[2] or 0) for r in data]

        bars = ax.bar(
            x,
            dt,
            width=0.55,
            color="#1976D2",
            edgecolor="white",
            linewidth=1.2
        )

        ymax = max(dt) if dt else 1

        ax.set_ylim(0, ymax * 1.2)

        for bar in bars:
            val = bar.get_height()

            ax.text(
                bar.get_x() + bar.get_width() / 2,
                val + ymax * 0.02,
                f"{val/1e6:.1f}M",
                ha="center",
                fontsize=8,
                fontweight="bold"
            )

        ax.set_title(
            "Doanh thu theo ngày",
            fontsize=16,
            fontweight="bold",
            pad=18
        )

    ax.set_xticks(x)

    ax.set_xticklabels(
        ngay,
        rotation=25 if n > 6 else 0,
        fontsize=9
    )

    ax.tick_params(axis="y", labelsize=9)

    ax.set_ylabel(
        "Số tiền (VNĐ)",
        fontsize=10,
        fontweight="bold"
    )

    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(
            lambda v, _: (
                f"{v/1e6:.1f}M"
                if abs(v) >= 1_000_000
                else f"{v/1e3:.0f}K"
            )
        )
    )

    ax.grid(
        axis="y",
        linestyle="--",
        alpha=0.3
    )

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_alpha(0.3)
    ax.spines["bottom"].set_alpha(0.3)

    fig.patch.set_facecolor("#F5F7FA")
    ax.set_facecolor("#FFFFFF")

    plt.tight_layout()

    return fig
